"""
******************************* CREDITSAFE CONNECTOR *********************************
CreditSafe ===========================================================================
Retrieves Company Information from Creditsafe and Updates Spreadsheet
Inputs:
 - Input Spreadsheet in SHEETIN or %1 with a specific format:
    - "Config" Sheet with all required fields
    - "Creditsafe" Sheet containing Company RegNos and Countries
 - Output Spreadsheet Template
 - CreditSafe: Correct Endpoints and Credentials in Spreadsheet above

Outputs:
 - Spreadsheet according to template, with one Sheet per Company
 - Codified Status in stdout
v1.1 (c) Delta AI 2022 - Diego Montoliu ==============================================
"""

# Dependencies ------------------------------------------------------------------------------------
import json
import requests
import openpyxl     # Excel Handling
from openpyxl.styles import Font
from sys import exit, argv
from os import path
# -------------------------------------------------------------------------------------------------


# Config ------------------------------------------------------------------------------------------
FILE_LOCATOR = path.dirname(path.realpath(argv[0])) + "\\"
SHEETIN = "Creditsafe.xlsm"

CS_AUTH_URL = "https://connect.creditsafe.com/v1/authenticate"  # Authentication EndPoint
CS_FIN_URL = "https://connect.creditsafe.com/v1/companies"      # Service EndPont
# -------------------------------------------------------------------------------------------------


# logea() -----------------------------------------------------------------------------------------
def logea(ptxt):
    """
    Writes to Creditsafe Log.txt and to stdout
    ptxt: Text to write
    """
    print(ptxt)
    with open(FILE_LOCATOR + "Creditsafe Log.txt", "a") as f:
        f.write(ptxt + "\n")
#  ---------------------------------------------------------------------------------------- logea()


# safeget() --------------------------------------------------------------------------------------
def safeget(pdic, ptag):
    """
    Returns a tag from a dictionary, or "" if empty
    pdic: Dictionary
    ptag: Key to retrieve
    returns: Retrieved value
    """
    if type(pdic) is not dict:
        return ""
    elif ptag == "":
        return ""
    elif ptag not in pdic.keys():
        return ""
    else:
        return pdic[ptag]
#  ------------------------------------------------------------------------------------- safeget()


# cs_readcfg() -----------------------------------------------------------------------------------
def cs_readcfg(psheet):
    """
    Reads Configuration from excel file
    sheet: Input spreadsheet
    returns: Configuration dictionary
    """
    assert psheet and len(psheet), "need a default worksheet name"
    lcscfg = dict()

    # Optionally, you can specify input sheet on the command line
    sheet = argv[1] if len(argv) > 1 else FILE_LOCATOR + psheet

    try:
        # Config is stored on fixed cells of the second tab of the input sheet
        lwb = openpyxl.load_workbook(sheet)
        sht = lwb.worksheets[1]
        lcscfg["user"] = sht.cell(2, 3).value       # C2: User
        lcscfg["passw"] = sht.cell(3, 3).value      # C3: Password
        lcscfg["authurl"] = sht.cell(5, 3).value    # C5: Authentication endpoint
        lcscfg["finurl"] = sht.cell(6, 3).value     # C6: Financial Info Endpoint
        lcscfg["country"] = sht.cell(8, 3).value    # C8: Default Country
        lcscfg["exefile"] = sht.cell(10, 3).value   # C10: Exe filename
        lcscfg["outfile"] = sht.cell(11, 3).value   # C11: Output filename
        lwb.close()
    except Exception as e:
        logea(f"[empty]")
        logea(f"Error Reading Sheet: {e}")
        exit(-1)
    return lcscfg
#  ---------------------------------------------------------------------------------- cs_readcfg()


# cs_readcomp() ----------------------------------------------------------------------------------
def cs_readcomp(psheet, pdefcountry):
    """
    Retrieves the companies to retrieve from input sheet
    psheet: Input spreadsheet
    defcountry: Default Country if empty
    returns: List of dictionaries with structure {'regno', 'country'}
    """
    assert psheet and len(psheet), "need a worksheet name"
    assert pdefcountry and len(pdefcountry), "need a default country"

    lcomps = []     # List containing company ids
    try:
        # Access the first tab of the input
        lwb = openpyxl.load_workbook(FILE_LOCATOR + psheet, data_only=True)
        sht = lwb.worksheets[0]

        # Iterate downwards starting from cell C2
        i = 3
        while sht.cell(i, 2).value:
            # Read id and country
            new_cmp = {'regno': sht.cell(i, 2).value, 'country': sht.cell(i, 4).value or pdefcountry}
            new_cmp['country'] = pdefcountry if new_cmp['country'][0] == "#" else new_cmp['country']
            lcomps.append(new_cmp)
            i += 1
        lwb.close()
    except Exception as e:
        logea(f"[empty]")
        logea(f"Error Reading Sheet: {e}")
        exit(-1)
    return lcomps
#  --------------------------------------------------------------------------------- cs_readcomp()


# cs_authenticate() ------------------------------------------------------------------------------
def cs_authenticate(purl, pusername, ppassword):
    """
    Authenticates using post data
    purl: Authentication Endpoint
    username: Auth. User
    password: Auth. Password
    returns: Authentication Token
    """
    assert purl and len(purl), "need an authentication url"
    assert pusername and len(pusername), "need an authentication username"
    assert ppassword and len(ppassword), "need an authentication password"

    tok = ""
    try:
        # Send a POST request with user/passwd
        postdata = {'username': pusername, 'password': ppassword}
        response = requests.post(url=purl, data=postdata)

        # Retrieve auth token
        tok = json.loads(response.text)
        tok = tok["token"] if "token" in tok.keys() else ""
    except ConnectionError as e:
        logea(f"[empty]")
        logea(f"Authentication connection failed: {e.strerror}")
        exit(-1)
    except Exception as e:
        logea(f"[empty]")
        logea(f"Authentication failed: {e}")
        exit(-1)
    return tok
#  ----------------------------------------------------------------------------- cs_authenticate()


# cs_compdata() ----------------------------------------------------------------------------------
def cs_compdata(ptoken, purl, pcompany):
    """
    Get CS Company Data
    ptoken: Authentication String
    url: Company Id Endpoint
    company: Regno to seek
    returns: Company CS Id
    """
    assert ptoken and len(ptoken), "ptoken must be a string"
    assert purl and len(purl), "need a company data url"
    assert pcompany and len(pcompany), "need a company"

    getdata = ({'countries': pcompany['country'], 'regNo': str(pcompany['regno']).strip()})
    compid = f"{pcompany['country']}-{pcompany['regno']}"
    try:
        # Retrieve Basic Company Data by providing country + RegNo
        response = requests.get(url=purl, params=getdata, headers={'Authorization': ptoken})
        if response.status_code in [400, 401, 403]:
            # Fail
            logea(f"[{compid}:{response.reason}]")
            logea(f"Company RegNo {pcompany['regno']} not found in country {pcompany['country']}. {response.reason}")
            lcompdata = ""
        else:
            # Success: Retrieve JSON
            lcompdata = json.loads(response.text)
            if "companies" in lcompdata.keys():
                lcompdata = lcompdata['companies'][0]
            else:
                logea(f"[{compid}:{lcompdata['details']}]")
                logea(f"Company RegNo {pcompany['regno']} not found in country {pcompany['country']}. {lcompdata['details']}")
                lcompdata = ""
    except Exception as e:
        logea(f"[{compid}:{e}]")
        logea(f"Company Id request failed: {e}")
        lcompdata = ""

    return lcompdata
#  --------------------------------------------------------------------------------- cs_compdata()


# cs_financialdata() -----------------------------------------------------------------------------
def cs_financialdata(ptoken, purl, pcompdata):
    """
    Get CS Financial Data
    ptoken: Authentication token
    url: Financial Data Endpoint
    pcompdata: Dict with the basic company info
    returns: JSONs with Financial Data, Local Financial Data, Group Structure, Employees
    """
    assert ptoken and len(ptoken), "ptoken must be a string"
    assert purl and len(purl), "url must be a string"
    assert pcompdata and len(pcompdata['id']), "missing company data"

    compid = f"{pcompdata['country']}-{pcompdata['regNo']}"
    try:
        # Retrieve Financial Data
        response = requests.get(url=f"{purl}/{pcompdata['id']}", headers={'Authorization': ptoken})
        if response.status_code in [400, 401, 403]:
            # Fail
            logea(f"[{compid}:{response.reason}]")
            logea(f"Failed to retrieve details: {response.reason}")
            fin_stat, local_fin_stat, grp_st, lemployees = "", "", "", ""
        else:
            # Success: Retrieve JSON
            rep = json.loads(response.text)
            rep = rep["report"] if "report" in rep.keys() else ""   # Data might be at top level or under "report"

            # Break down information
            fin_stat = safeget(rep, "financialStatements")
            local_fin_stat = safeget(rep, "localFinancialStatements")
            grp_st = safeget(rep, "groupStructure")
            lemployees = safeget(rep, "otherInformation")
            lemployees = safeget(lemployees, "employeesInformation")
    except ConnectionError as e:
        logea(f"[{compid}:{e.strerror}]")
        logea(f"Financial Data connection failed: {e.strerror}")
        fin_stat, local_fin_stat, grp_st, lemployees = "", "", "", ""
    except Exception as e:
        logea(f"[{compid}:{e}]")
        logea(f"Parsing failed: {e}")
        fin_stat, local_fin_stat, grp_st, lemployees = "", "", "", ""

    return fin_stat, local_fin_stat, grp_st, lemployees
#  ---------------------------------------------------------------------------- cs_financialdata()


# cs_cleansheet() --------------------------------------------------------------------------------
def cs_cleansheet(psheet):
    """
    Cleans old tabs from Results
    sheet: Results Sheet path
    returns: Results Sheet object
    """
    assert psheet and len(psheet), "need an output spreadsheet"

    try:
        # Remove all sheet except the Template one
        lwb = openpyxl.load_workbook(FILE_LOCATOR + psheet)
        while len(lwb.worksheets) > 1:
            sht = lwb.worksheets[-1]
            lwb.remove(sht)
    except Exception as e:
        logea("[empty]")
        logea(f"Error Updating Sheet: {e}")
        lwb = None
    return lwb
#  -------------------------------------------------------------------------------- cs_cleansheet()


# cs_updatesheet() -------------------------------------------------------------------------------
def cs_updatesheet(pcompdata, pfinstat, plocfinstat, pgrpstruc, pemployees, pwb):
    """
    Updates Output sheet with the retrieved info
    pcompdata: Company JSON
    pfinstat: Financial Data JSON
    plocfinstat: Local Financial Data JSON
    pgrpstruc: Group Structure JSON
    pemployees: Employees JSON
    pwb: Output Excel File Object
    returns: Nothing
    """
    assert pcompdata and isinstance(pcompdata, dict), "no company data"
    assert pfinstat and isinstance(pfinstat, list), "no financial data"
    assert plocfinstat and isinstance(plocfinstat, list), "no local financial data"
    assert pemployees and isinstance(pemployees, list), "no employee data"
    assert pwb and len(pwb.worksheets), "incorrect results excel file"

    # Finds a specific tag and year in the Financial series
    def finentry(pi, tag):
        if pi < len(pfinstat) and tag in pfinstat[pi].keys():
            return pfinstat[pi][tag]
        elif pi < len(plocfinstat) and tag in plocfinstat[pi].keys():
            return plocfinstat[pi][tag]
        elif pi < len(pfinstat) and 'profitAndLoss' in pfinstat[pi].keys() and tag in pfinstat[pi][
                'profitAndLoss'].keys():
            return pfinstat[pi]['profitAndLoss'][tag]
        elif pi < len(plocfinstat) and 'profitAndLoss' in plocfinstat[pi].keys() and tag in plocfinstat[pi]['profitAndLoss'].keys():
            return plocfinstat[pi]['profitAndLoss'][tag]
        else:
            return ""

    # Finds a specific year in the Employees series
    def empentry(pi):
        if pi >= len(pemployees):
            return ""
        elif pemployees[pi] == "":
            return ""
        elif "numberOfEmployees" in pemployees[pi].keys():
            return pemployees[pi]["numberOfEmployees"]
        else:
            return ""

    # Returns the long description of the Company
    def showcmp(tag):
        if pgrpstruc == "":
            return "n/a"
        elif tag in pgrpstruc.keys():
            country = safeget(pgrpstruc[tag], "country")
            lregno = safeget(pgrpstruc[tag], "registrationNumber")
            return f'{pgrpstruc[tag]["name"]} ({country} - {lregno})'
        else:
            return "n/a"

    # Show the Group Structure, "tag" subtree, on row prow
    def showcmplst(tag, prow):
        new_row = prow
        if pgrpstruc == "":
            return new_row
        if tag in pgrpstruc.keys():
            for cmp in pgrpstruc[tag]:
                country = safeget(cmp, "country")
                lregno = safeget(cmp, "registrationNumber")
                sht.cell(new_row, 11).value = f'{cmp["name"]} ({country} - {lregno})'
                new_row += 1
                if new_row - prow > 10:
                    break
        return new_row

    # Update, main loop
    compid = f"{pcompdata['country']}-{pcompdata['regNo']}"
    try:
        # Create a new sheet with the company name
        sht = pwb.copy_worksheet(pwb.worksheets[0])
        sht.cell(14, 4).value = pcompdata['name'] if "name" in pcompdata.keys() else "n/a"
        sht.title = compid

        # Loop over last five financial years
        for i in range(5):
            # Loop over rows. Fill cells according to the tag on the "A" Column until cell is empty
            col = 8-i
            row = 1
            cll = sht.cell(row, 1).value
            while cll:
                if cll == "#employees":
                    sht.cell(row, col).value = empentry(i)
                elif cll[0] != "#":
                    sht.cell(row, col).value = finentry(i, cll)
                row += 1
                cll = sht.cell(row, 1).value

        # Fill Group Structure Starting from K14 downwards
        sht.cell(14, 11).value = showcmp("ultimateParent")
        sht.cell(15, 11).value = showcmp("immediateParent")
        row = 16
        row = showcmplst("subsidiaryCompanies", row)
        sht.cell(row, 10).value = "Affiliates"
        sht.cell(row, 10).font = Font(name='Arial', size=14)
        showcmplst("affiliatedCompanies", row)
        logea(f"[{sht.title}:OK]")
    except Exception as e:
        logea(f"[{compid}:{e}]")
        logea(f"Error Updating Sheet: {e}")
        exit(-1)
    return
#  ------------------------------------------------------------------------------ cs_updatesheet()


# MAIN -------------------------------------------------------------------------------------------
logea(f"CreditSafe Armstrong Connector v1.1 (c) 2022 Delta AI")
cscfg = cs_readcfg(SHEETIN)
comps = cs_readcomp(SHEETIN, cscfg["country"])

# Check if anything to process
if comps is None or len(comps) == 0:
    logea("[empty]")
    logea(f"No Companies to Retrieve")
    exit(0)

token = cs_authenticate(cscfg["authurl"], cscfg["user"], cscfg["passw"])        # Retrieve authentication ptoken
if len(token) == 0:
    logea("[empty]")
    logea(f"Login to CreditSafe Failed")
    exit(-1)

logea(f"Connected to CreditSafe")
wb = cs_cleansheet(cscfg["outfile"])  # Clean Spreadsheet
if wb is None:  # Fail
    logea("[empty]")
    logea(f"Could not access {cscfg['outfile']}")
    exit(-1)

# Loop through companies and fill in one tab per company
for comp in comps:
    compdata = cs_compdata(token, cscfg["finurl"], comp)      # Retrieve Company Id
    if len(compdata):
        finstat, locfinstat, grpstruc, employees = cs_financialdata(token, cscfg["finurl"], compdata)  # Retrieve Financial Data
        if len(finstat):
            cs_updatesheet(compdata, finstat, locfinstat, grpstruc, employees, wb)                         # Update Spreadsheet
            logea(f"Successfully Updated {cscfg['outfile']} with {comp}")
    else:
        logea(f"Could not update {cscfg['outfile']} with {comp}")

# Save file and terminate
if len(wb.worksheets) > 1:
    wb.save(FILE_LOCATOR + cscfg["outfile"])
    logea(f"[OK]")
    logea(f"Finished")
else:
    logea(f"[empty]")
    logea(f"Nothing Retrieved")
#  ------------------------------------------------------------------------------------------ MAIN
